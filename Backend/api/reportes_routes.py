# api/reportes_routes.py

from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt

from io import BytesIO
from datetime import datetime

from core.extensions import db
from core.auth import role_required

from models.ordenes import OrdenTrabajo
from models.personas import Cliente, Empleado
from models.vehiculos import Motocicleta
from models.reportes import ReporteTrabajo  # <-- este es tu modelo real

# para XLSX
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# para PDF
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors


reportes_bp = Blueprint('reportes', __name__, url_prefix='/api/reportes')


def _current_user():
    """
    Saca info básica del usuario autenticado desde el JWT.
    Espera que en el token guardaste algo como:
    {
        "user": {
            "id": 1,
            "usuario": "admin",
            "rol": "gerente"
        }
    }
    """
    claims = get_jwt()
    return claims.get("user", {})


# --------------------------------------------------------------------
# 1) DATASET GENERAL (para XLSX/PDF global o por cliente)
# --------------------------------------------------------------------
def _armar_dataset(cliente_id=None):
    """
    Arma la data cruda que vamos a exportar.
    Si cliente_id viene, filtramos solo ese cliente.
    Si no viene, sacamos todas las órdenes.

    Retorna una lista de dicts, cada dict = UNA orden con:
    - datos de la orden
    - lista de trabajos realizados (reportes_trabajo) en texto formateado
    """

    # query base de órdenes con joins
    q = (
        OrdenTrabajo.query
        .join(Cliente, Cliente.id == OrdenTrabajo.cliente_id)
        .join(Motocicleta, Motocicleta.id == OrdenTrabajo.motocicleta_id)
        .outerjoin(Empleado, Empleado.id == OrdenTrabajo.mecanico_asignado_id)
        .order_by(OrdenTrabajo.fecha_ingreso.desc())
    )

    if cliente_id:
        q = q.filter(OrdenTrabajo.cliente_id == cliente_id)

    ordenes = q.all()

    # Para cada orden también traemos sus reportes técnicos (ReporteTrabajo)
    # y los armamos en formato tipo:
    # "[31/10/2025 02:39, Edinilson Valdes] cambio de aceite"
    rows = []
    for o in ordenes:
        cliente = getattr(o, "cliente", None)
        moto = getattr(o, "motocicleta", None)
        mecanico = getattr(o, "mecanico_asignado", None)

        # marca / modelo vienen anidados en motocicleta.modelo / modelo.marca
        marca_nombre = None
        modelo_nombre = None
        if moto and getattr(moto, "modelo", None):
            modelo_nombre = moto.modelo.nombre
            if getattr(moto.modelo, "marca", None):
                marca_nombre = moto.modelo.marca.nombre

        # buscar los reportes_trabajo de ESTA orden
        # orden_id -> o.id
        reportes = (
            ReporteTrabajo.query
            .filter(ReporteTrabajo.orden_id == o.id)
            .order_by(ReporteTrabajo.creado_en.asc())
            .all()
        )

        trabajos_list = []
        for rep in reportes:
            fecha_txt = rep.creado_en.strftime("%d/%m/%Y %H:%M") if rep.creado_en else "¿?"
            linea = f"[{fecha_txt}, {rep.mecanico_nombre or '—'}] {rep.descripcion or ''}"
            trabajos_list.append(linea)

        trabajos_txt = "\n".join(trabajos_list)

        rows.append({
            "orden_id": o.id,
            "fecha_ingreso": o.fecha_ingreso.isoformat() if o.fecha_ingreso else None,
            "fecha_salida": o.fecha_salida.isoformat() if getattr(o, "fecha_salida", None) else None,
            "estado": o.estado.value if o.estado else None,

            "cliente_nombre": cliente.nombre if cliente else None,
            "cliente_telefono": cliente.telefono if cliente else None,

            "moto_placa": moto.placa if moto else None,
            "moto_vin": moto.vin if moto else None,
            "moto_marca": marca_nombre,
            "moto_modelo": modelo_nombre,

            "mecanico_nombre": mecanico.nombre if mecanico else "Sin asignar",
            "observaciones": o.observaciones or "",

            # <-- lo nuevo:
            "trabajos_realizados": trabajos_txt,
        })

    return rows


# --------------------------------------------------------------------
# 2) EXPORT GLOBAL XLSX
# --------------------------------------------------------------------
def _export_xlsx(rows, titulo="Reporte de Órdenes"):
    """
    Genera un XLSX en memoria y lo devuelve como BytesIO listo para send_file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Ordenes"

    # encabezado
    headers = [
        "ID Orden",
        "Fecha Ingreso",
        "Fecha Salida",
        "Estado",
        "Cliente",
        "Teléfono",
        "Placa",
        "VIN",
        "Marca",
        "Modelo",
        "Mecánico Asignado",
        "Observaciones",
        "Trabajos realizados",
    ]
    ws.append(headers)

    # filas
    for r in rows:
        ws.append([
            r["orden_id"],
            r["fecha_ingreso"],
            r["fecha_salida"],
            r["estado"],
            r["cliente_nombre"],
            r["cliente_telefono"],
            r["moto_placa"],
            r["moto_vin"],
            r["moto_marca"],
            r["moto_modelo"],
            r["mecanico_nombre"],
            r["observaciones"],
            r["trabajos_realizados"],
        ])

    # ancho automático decente
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(18, len(header) + 2)

    # guardarlo en bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{titulo.replace(' ', '_').lower()}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return output, filename, mimetype


# --------------------------------------------------------------------
# 3) EXPORT GLOBAL PDF BONITO (tarjetas por orden)
# --------------------------------------------------------------------
def _export_pdf(rows, titulo="Reporte de Órdenes"):
    """
    Genera un PDF presentable:
    - Cada orden es una 'tarjeta' con borde gris y padding
    - Etiquetas alineadas
    - Trabajos realizados tipo lista con viñitas
    """

    buffer = BytesIO()

    page_width, page_height = A4
    c = canvas.Canvas(buffer, pagesize=A4)

    margin_x = 2 * cm
    margin_y = 2 * cm

    # posición inicial arriba de la página
    y = page_height - margin_y

    # Título grande
    c.setFont("Helvetica-Bold", 16)
    c.drawString(margin_x, y, titulo)
    y -= 1 * cm

    # estilos
    label_font = ("Helvetica-Bold", 9)
    value_font = ("Helvetica", 9)

    card_padding_x = 0.6 * cm
    card_padding_y = 0.5 * cm
    card_spacing_y = 0.8 * cm  # espacio entre tarjetas
    card_border_color = colors.grey

    line_height = 0.45 * cm

    def draw_label_value(label, value, cur_x, cur_y):
        """Dibuja `Label: value` alineado bonito"""
        c.setFont(*label_font)
        c.drawString(cur_x, cur_y, f"{label}:")
        label_w = c.stringWidth(f"{label}:", label_font[0], label_font[1])
        c.setFont(*value_font)
        c.drawString(cur_x + label_w + 4, cur_y, value if value else "—")

    for r in rows:
        # formatear la lista de trabajos
        trabajos_bloque = []
        trabajos_txt = r.get("trabajos_realizados", "") or ""
        if trabajos_txt.strip():
            for raw_line in trabajos_txt.split("\n"):
                trabajos_bloque.append("• " + raw_line.strip())

        # Estimación de alto de tarjeta
        base_lines = 7  # encabezado + fechas + cliente + moto + modelo + mecánico + obs + título trabajos
        trabajos_lines = len(trabajos_bloque) if trabajos_bloque else 1
        total_lines = base_lines + trabajos_lines

        card_height = (total_lines * line_height) + (2 * card_padding_y)

        # Salto de página si no cabe
        if y - card_height < margin_y:
            c.showPage()
            page_width, page_height = A4
            y = page_height - margin_y

            c.setFont("Helvetica-Bold", 16)
            c.drawString(margin_x, y, titulo)
            y -= 1 * cm

        # coords tarjeta
        card_x1 = margin_x
        card_y1 = y - card_height
        card_x2 = page_width - margin_x
        card_y2 = y

        # borde tarjeta
        c.setStrokeColor(card_border_color)
        c.setLineWidth(0.5)
        c.roundRect(card_x1, card_y1, card_x2 - card_x1, card_y2 - card_y1, 6, stroke=1, fill=0)

        # contenido
        text_x = card_x1 + card_padding_x
        text_y = card_y2 - card_padding_y

        # Encabezado orden + estado
        c.setFont("Helvetica-Bold", 11)
        encabezado = f"Orden #{r['orden_id']}   |   {r['estado'] or 'SIN ESTADO'}"
        c.drawString(text_x, text_y, encabezado)
        text_y -= line_height

        # Fechas
        c.setFont("Helvetica", 9)
        fechas_line = (
            f"Ingreso: {r['fecha_ingreso'] or '—'}    "
            f"Salida: {r['fecha_salida'] or '—'}"
        )
        c.drawString(text_x, text_y, fechas_line)
        text_y -= line_height

        # Cliente
        draw_label_value(
            "Cliente",
            f"{r['cliente_nombre'] or '—'}   Tel: {r['cliente_telefono'] or '—'}",
            text_x,
            text_y
        )
        text_y -= line_height

        # Moto
        draw_label_value(
            "Motocicleta",
            f"Placa {r['moto_placa'] or '—'} / VIN {r['moto_vin'] or '—'}",
            text_x,
            text_y
        )
        text_y -= line_height

        # Modelo
        draw_label_value(
            "Modelo",
            f"{r['moto_marca'] or '—'} {r['moto_modelo'] or ''}".strip(),
            text_x,
            text_y
        )
        text_y -= line_height

        # Mecánico asignado
        draw_label_value(
            "Mecánico",
            r['mecanico_nombre'] or '—',
            text_x,
            text_y
        )
        text_y -= line_height

        # Observaciones
        draw_label_value(
            "Observaciones",
            r['observaciones'] or '—',
            text_x,
            text_y
        )
        text_y -= line_height

        # línea separadora fina
        c.setStrokeColor(colors.lightgrey)
        c.setLineWidth(0.3)
        c.line(
            text_x,
            text_y + (line_height * 0.4),
            card_x2 - card_padding_x,
            text_y + (line_height * 0.4)
        )
        text_y -= (line_height * 0.6)

        # Trabajos realizados
        c.setFont("Helvetica-Bold", 9)
        c.drawString(text_x, text_y, "Trabajos realizados:")
        text_y -= line_height

        c.setFont("Helvetica", 9)

        if trabajos_bloque:
            for linea_rep in trabajos_bloque:
                # wrap manual para no salirnos de la tarjeta
                max_width = (card_x2 - card_padding_x) - text_x
                words = linea_rep.split()
                actual = ""
                for w in words:
                    probe = (actual + " " + w).strip()
                    if c.stringWidth(probe, "Helvetica", 9) > max_width:
                        c.drawString(text_x, text_y, actual)
                        text_y -= line_height
                        actual = w
                    else:
                        actual = probe
                if actual:
                    c.drawString(text_x, text_y, actual)
                    text_y -= line_height
        else:
            c.drawString(text_x, text_y, "—")
            text_y -= line_height

        # bajar Y para siguiente tarjeta
        y = card_y1 - card_spacing_y

    # cerrar PDF
    c.showPage()
    c.save()

    buffer.seek(0)

    filename = f"{titulo.replace(' ', '_').lower()}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
    mimetype = "application/pdf"
    return buffer, filename, mimetype


# --------------------------------------------------------------------
# 4) ENDPOINT: EXPORT ORDENES (GLOBAL / POR CLIENTE)
# --------------------------------------------------------------------
@reportes_bp.route('/ordenes', methods=['GET'])
@jwt_required()
@role_required("gerente", "encargado")
def exportar_ordenes():
    """
    Exporta órdenes de trabajo completas.
    Query params:
      - formato = 'xlsx' | 'pdf'   (obligatorio)
      - cliente_id (opcional) -> si viene, filtra solo ese cliente

    Ejemplos desde el front:
      /api/reportes/ordenes?formato=xlsx
      /api/reportes/ordenes?formato=pdf
      /api/reportes/ordenes?formato=pdf&cliente_id=7
    """

    formato = (request.args.get('formato') or '').lower().strip()
    cliente_id = request.args.get('cliente_id', type=int)

    if formato not in ('xlsx', 'pdf'):
        return jsonify({"error": "formato inválido, use xlsx o pdf"}), 400

    rows = _armar_dataset(cliente_id=cliente_id)

    # si filtra por cliente_id, cambiamos título para que el archivo se llame bonito
    if cliente_id:
        titulo = f"Historial Cliente #{cliente_id}"
    else:
        titulo = "Ordenes Taller"

    if formato == 'xlsx':
        output, filename, mimetype = _export_xlsx(rows, titulo=titulo)
    else:
        output, filename, mimetype = _export_pdf(rows, titulo=titulo)

    return send_file(
        output,
        mimetype=mimetype,
        as_attachment=True,
        download_name=filename
    )